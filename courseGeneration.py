import pandas as pd
import os
import random
import openpyxl
from openpyxl.styles import Font, Alignment

# test cases: max is 1000 students from student.xlsx. Random student-course is generated to studentCourses.xlsx
NUM_STUDENTS = 800


file_path_students_names = '/Users/ez/Documents/Course-Sorting-Algorithm/exampleInput/student.xlsx'
if not os.path.exists(file_path_students_names):
    print(f"Error: The file {file_path_students_names} does not exist.")
    exit()

df_student_names = pd.read_excel(file_path_students_names)
studentsNames = df_student_names.iloc[:, 0].tolist()
selected_students = studentsNames[:NUM_STUDENTS]

student_dict = {name: 218620 + index for index, name in enumerate(selected_students)}

print(f"Number of students processed: {len(student_dict)}")
print(f"Sample student data: {list(student_dict.items())[:5]}")


# Required courses for each grade: In the future this must not be hard Coded in.
grade_8_required = [("ENGLISH 8", 8),
                    ("MATHEMATICS 8", 8),
                    ("FRENCH 8", 8),
                    ("ADST Rotation", 8),
                    ("Fine Arts Rotation", 8),
                    ("SOCIAL STUDIES 8", 8),
                    ("PHYSICAL AND HEALTH EDUCATION 8", 8),
                    ("SCIENCE 8", 8)
                    ]

grade_9_required = [("ENGLISH 9", 9),
                    ("MATHEMATICS 9", 9),
                    ("SOCIAL STUDIES 9", 9),
                    ("PHYSICAL AND HEALTH EDUCATION 9", 9),
                    ("SCIENCE 9", 9)
                    ]

grade_10_required = [("ENGLISH 10", 10),
                     ("FOUNDATION OF MATHEMATICS & PRE-CALCULUS 10", 10),
                     ("SOCIAL STUDIES 10", 10),
                     ("PHYSICAL AND HEALTH EDUCATION 10", 10),
                     ("SCIENCE 10", 10),
                     ("CAREER AND LIFE EXPLORATION 10 CLE", 10)
                     ]

grade_11_required = [("FOCUSED LITERARY STUDIES 11", 11),
                     ("PRE-CALCULUS 11", 11),
                     ("SOCIAL STUDIES 11", 11),
                     ]

grade_12_required = [("ENGLISH STUDIES 12", 12),
                     ("PRE-CALCULUS 12", 12),
                     ("CAREER & LIFE CONNECTIONS (CLC) & CAPSTONE", 12)
                     ]

language_courses = [("FRENCH 9", 9),
                    ("FRENCH 10", 10),
                    ("FRENCH 11", 11),
                    ("FRENCH 12", 12),
                    ("SPANISH 10", 10),
                    ("SPANISH 11", 11),
                    ("SPANISH 12", 12)
                    ]

adst_courses = [("FOOD STUDIES 9", 9), ("FOOD STUDIES 10 & INTRO 11", 10),
                ("FOOD STUDIES 11", 11),
                ("FOOD STUDIES 12", 12),
                ("INFORMATION & COMMUNICATIONS TECHNOLOGIES 9", 9),
                ("COMPUTER STUDIES 10", 10),
                ("COMPUTER INFORMATION SYSTEMS 11", 11),
                ("COMPUTER INFORMATION SYSTEMS 12", 12),
                ("COMPUTER PROGRAMMING 12", 12),
                ("DRAFTING 11", 11),
                ("DRAFTING 12", 12),
                ("ENGINEERING 11", 11),
                ("ENGINEERING 12", 12)
                ]
fine_arts_courses = [("MUSIC 9: CONCERT CHOIR", 9),
                     ("CHORAL MUSIC 10: CONCERT CHOIR", 10),
                     ("CHORAL MUSIC 11: CONCERT CHOIR", 11),
                     ("CHORAL MUSIC 12: CONCERT CHOIR", 12),
                     ("DRAMA 9 (ACTING)", 9), ("DRAMA 10 (ACTING)", 10),
                     ("DRAMA 11 (ACTING)", 11), ("DRAMA 12 (ACTING)", 12),
                     ("VISUAL ARTS 9 (DRAWING & PAINTING)", 9),
                     ("STUDIO ARTS 2D 10 (DRAWING & PAINTING)", 10),
                     ("STUDIO ARTS 2D 11 (DRAWING & PAINTING)", 11),
                     ("STUDIO ARTS 2D 12 (DRAWING & PAINTING)", 12)
                     ]
science_11_12 = [("LIFE SCIENCES 11", 11),
                 ("ANATOMY AND PHYSIOLOGY 12", 12),
                 ("CHEMISTRY 11", 11),
                 ("CHEMISTRY 12", 12),
                 ("PHYSICS 11", 11),
                 ("PHYSICS 12", 12)
                 ]
grade_12_electives = [("AP MICROECONOMICS 12", 12),
                      ("AP ENGLISH LITERATURE AND COMPOSITION 12", 12),
                      ("AP CALCULUS 12 AB", 12),
                      ("AP BIOLOGY 12", 12),
                      ("AP CHEMISTRY 12", 12),
                      ("AP PHYSICS 2 HONOURS 12", 12),
                      ("AP HUMAN GEOGRAPHY 12", 12)
                      ]

def generate_course_selection(grade):
    courses = []
    try:
        if grade == 8:
            courses = [course for course, _ in grade_8_required]
        elif grade == 9:
            courses = [course for course, _ in grade_9_required]
            courses.append(
                random.choice([course for course, course_grade in language_courses if course_grade == grade]))
            courses.append(random.choice([course for course, course_grade in adst_courses if course_grade == grade]))
            courses.append(
                random.choice([course for course, course_grade in fine_arts_courses if course_grade == grade]))
        elif grade == 10:
            courses = [course for course, _ in grade_10_required]
            courses.append(
                random.choice([course for course, course_grade in language_courses if course_grade == grade]))
            adst_fine_arts_choice = random.choice(
                [course for course, course_grade in adst_courses + fine_arts_courses if course_grade == grade])
            courses.append(adst_fine_arts_choice)
        elif grade == 11:
            courses = [course for course, _ in grade_11_required]
            courses.extend(
                random.sample([course for course, course_grade in science_11_12 if course_grade == grade], 2))
            courses.append(random.choice([course for course, course_grade in adst_courses if course_grade == grade]))
            courses.append(
                random.choice([course for course, course_grade in fine_arts_courses if course_grade == grade]))
            # Add one more elective to reach 8 courses
            additional_elective = random.choice(
                [course for course, course_grade in language_courses + adst_courses + fine_arts_courses if
                 course_grade == grade])
            courses.append(additional_elective)
        elif grade == 12:
            courses = [course for course, _ in grade_12_required]
            courses.extend(
                random.sample([course for course, course_grade in science_11_12 if course_grade == grade], 2))
            courses.append(random.choice([course for course, _ in grade_12_electives]))
            adst_fine_arts_choices = random.sample(
                [course for course, course_grade in adst_courses + fine_arts_courses if course_grade == grade], 2)
            courses.extend(adst_fine_arts_choices)
        else:
            raise ValueError(f"Invalid grade: {grade}")

        if len(courses) != 8:
            raise ValueError(f"Incorrect number of courses generated for grade {grade}: {len(courses)}")
        return courses
    except Exception as e:
        print(f"Error generating courses for grade {grade}: {str(e)}")
        return []

def generate_preferences(grade):
    preferences = []
    try:
        if grade in [9, 10, 11, 12]:
            # Filter ADST and Fine Arts courses by the student's grade
            adst_choices = [course for course, course_grade in adst_courses if course_grade == grade]
            fine_arts_choices = [course for course, course_grade in fine_arts_courses if course_grade == grade]

            # Ensure there are enough courses to sample from
            adst_sample_size = min(3, len(adst_choices))
            fine_arts_sample_size = min(2, len(fine_arts_choices))

            preferences += random.sample(adst_choices, adst_sample_size)
            preferences += random.sample(fine_arts_choices, fine_arts_sample_size)

        return preferences[:5]  # Return top 5 preferences
    except Exception as e:
        print(f"Error generating preferences for grade {grade}: {str(e)}")
        return []

# Create students dictionary with course selections and preferences
students = {}
for name, number in student_dict.items():
    grade = random.choice([8, 9, 10, 11, 12])
    students[name] = {
        "number": number,
        "grade": grade,
        "courses": generate_course_selection(grade),
        "preferences": generate_preferences(grade)
    }

def export_to_excel(students, filename='studentCourses.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Courses"

    headers = ["Student Name", "Student Number", "Grade", "Courses", "Preferences"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row, (name, data) in enumerate(students.items(), start=2):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=data['number'])
        ws.cell(row=row, column=3, value=data['grade'])
        ws.cell(row=row, column=4, value=', '.join(data['courses']))
        ws.cell(row=row, column=5, value=', '.join(data['preferences']))  # Changed line

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(path/to/Studentname)
    print(f"Data exported to {filename}")

export_to_excel(students)
